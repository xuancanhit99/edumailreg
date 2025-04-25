from PyQt6.uic import loadUi
from PyQt6.QtWidgets import QApplication, QMainWindow, QMessageBox
from PyQt6.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook, Workbook
from modules.Bot import Bot
from modules.ChangePasswordBot import ChangePasswordBot
import sys
import os


class WorkerThread(QThread):
    finished = pyqtSignal()

    def __init__(self, config, log_callback):
        super().__init__()
        self.config = config
        self.add_log = log_callback
        self.is_running = False

        self.headers = ["Email", "Student ID", "Password", "Full Name", "Gender",
                        "Birthdate", "Street", "City", "State", "Zipcode", "SSN", "Password changed"]

    def stop(self):
        self.is_running = False

    def run(self):
        self.is_running = True

        data_file = self.config['data_file']
        backup_file = self.config['backup_file']

        if os.path.exists(data_file):
            workbook = load_workbook(data_file)
            sheet = workbook.active

            if self.config['backup_enabled']:
                workbook.save(backup_file)
                self.add_log(f"Đã tạo file backup: {backup_file}")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(self.headers)

        if self.config['mail_enabled'] and self.config['password_enabled']:
            self.run_create_and_change_password(sheet, workbook)
        elif self.config['mail_enabled']:
            self.run_create_only(sheet, workbook)
        elif self.config['password_enabled']:
            self.run_change_existing_passwords(sheet, workbook)

        self.finished.emit()

    def run_create_and_change_password(self, sheet, workbook):
        bot = Bot(token=self.config['token'], wait_sec=self.config['wait_sec'],
                  headless_mode=self.config['headless_mode'])
        change_pw_bot = ChangePasswordBot(token=self.config['token'])
        success_count = 0

        for i in range(1, self.config['tries'] + 1):
            if not self.is_running:
                self.add_log("💀 Chương trình đã bị dừng bởi người dùng.")
                break

            result = bot.run()
            
            if not result:
                self.add_log(f"❌ Có lỗi xảy ra khi tạo tài khoản {i}")
                continue

            self.add_log(f"✅ Tài khoản {i} tạo thành công: {result.get('Student ID')}")

            new_password = (change_pw_bot.generate_password() if self.config['create_new_password']
                            else result.get("Password"))

            try:
                change_result = change_pw_bot.run(
                    email=result["Email"],
                    sfid=result["Student ID"],
                    password=new_password,
                    wait_sec_get_mail=self.config['wait_sec_get_mail']
                )

                if change_result:
                    self.add_log(f"✅ Đổi thành công mật khẩu cho {result['Student ID']}")
                    result["Password"] = change_result
                    result["Password changed"] = "y"

            except Exception as e:
                self.add_log(f"❌ Có lỗi xảy ra khi đổi mật khẩu cho id {result['Student ID']}: {e}")

            sheet.append([result.get(h, "") for h in self.headers])
            workbook.save(self.config['data_file'])
            success_count += 1
            self.add_log(f"🚩 Đã tạo {success_count}/{self.config['tries']} tài khoản")

    def run_create_only(self, sheet, workbook):
        bot = Bot(token=self.config['token'], wait_sec=self.config['wait_sec'],
                  headless_mode=self.config['headless_mode'])
        success_count = 0

        for i in range(1, self.config['tries'] + 1):
            if not self.is_running:
                self.add_log("💀 Chương trình đã bị dừng bởi người dùng.")
                break
            
            result = bot.run()

            if not result:
                self.add_log(f"❌ Đã có lỗi xảy ra khi tạo tài khoản {i}")
                continue

            sheet.append([result.get(h, "") for h in self.headers])
            workbook.save(self.config['data_file'])
            success_count += 1
            self.add_log(f"✅ Tài khoản {i} tạo thành công: {result.get('Student ID')}")
            self.add_log(f"🚩 Đã tạo {success_count}/{self.config['tries']} tài khoản")

    def run_change_existing_passwords(self, sheet, workbook):
        change_pw_bot = ChangePasswordBot(token=self.config['token'])
        success_count = 0

        for row in sheet.iter_rows(min_row=2):
            if not self.is_running:
                self.add_log("💀 Chương trình đã bị dừng bởi người dùng.")
                break

            if row[11].value == "n":
                email, sfid, current_pw = row[0].value, row[1].value, row[2].value
                new_password = (change_pw_bot.generate_password() if self.config['create_new_password'] else current_pw)

                try:
                    change_result = change_pw_bot.run(
                        email=email,
                        sfid=sfid,
                        password=new_password,
                        wait_sec_get_mail=self.config['wait_sec_get_mail']
                    )

                    if change_result:
                        row[2].value = change_result
                        row[11].value = "y"
                        workbook.save(self.config['data_file'])
                        success_count += 1
                        self.add_log(f"✅ Đã đổi mật khẩu cho id {sfid}")
                except Exception as e:
                    self.add_log(f"❌ Có lỗi xảy ra khi đổi mật khẩu cho {sfid}: {e}")

        self.add_log(f"🚩 Đã đổi thành công mật khẩu cho {success_count} tài khoản.")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        loadUi('ui/main.ui', self)

        self.run.clicked.connect(self.start_thread)
        self.stop.clicked.connect(self.stop_thread)

        self.mail.toggled.connect(self.update_ui_state)
        self.password.toggled.connect(self.update_ui_state)

        self.token_file = "token.txt"
        self.saveToken.clicked.connect(self.save_token_to_file)
        self.load_token()

        self.thread = None
        self.update_ui_state()

    def save_token_to_file(self):
        with open(self.token_file, 'w') as file:
            file.write(self.tempMailToken.text())
            self.log.addItem(f"Đã lưu token vào file: {self.token_file}")
    
    def load_token(self):
        if os.path.exists(self.token_file):
            with open(self.token_file, "r") as file:
                self.tempMailToken.setText(file.read())

    def update_ui_state(self):
        can_run = self.mail.isChecked() or self.password.isChecked()
        self.run.setEnabled(can_run and not self.thread)
        self.stop.setEnabled(bool(self.thread))

    def start_thread(self):
        token = self.tempMailToken.text()
        data_file = self.dataFile.text()
        backup_file = self.backupFile.text()

        if not token:
            self.show_message("Nhập token của TempMail trước khi chạy.")
            return

        if not data_file:
            self.show_message("Nhập tên file dữ liệu trước khi chạy.")
            return

        if self.backupFile.isEnabled() and not backup_file:
            self.show_message("Nhập tên file backup trước khi chạy.")
            return
            
        config = {
            'token': token,
            'tries': self.quantity.value(),
            'wait_sec': self.waitTime.value(),
            'headless_mode': self.headlessMode.isChecked(),
            'data_file': data_file,
            'backup_file': backup_file,
            'backup_enabled': self.isBackup.isChecked(),
            'create_new_password': self.newPass.isChecked(),
            'wait_sec_get_mail': self.waitSecGetMail.value(),
            'mail_enabled': self.mail.isChecked(),
            'password_enabled': self.password.isChecked()
        }

        self.thread = WorkerThread(config, self.log.addItem)
        self.thread.finished.connect(self.thread_finished)
        self.thread.start()
        self.update_ui_state()

    def thread_finished(self):
        self.thread = None
        self.update_ui_state()

    def stop_thread(self):
        if self.thread:
            self.thread.stop()

    def show_message(self, text, title="Warning"):
        QMessageBox.warning(self, title, text)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("WindowsVista")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
